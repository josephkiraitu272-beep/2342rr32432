"""
Guru — backend ingest API for Rozetka & Epicentr Chrome Extension.

Schema v1.0 (frozen). New fields are additive; existing fields keep their semantics.

Collections:
  products         — latest captured product snapshot per ingest event
  pages            — page-level summaries
  cf_sessions      — Cloudflare cookies forwarded from extension
  price_snapshots  — immutable price-history events keyed by product_key
"""
from fastapi import FastAPI, APIRouter, HTTPException, Header, Depends, Query
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import csv
import io
import hashlib
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Literal, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from urllib.parse import urlparse

from services.signals import build_signals_for_ingest
from services.intelligence_api import make_router as make_intelligence_router


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

EXTENSION_API_KEY = os.environ.get("EXTENSION_API_KEY", "prizma_dev_key_2026")
SCHEMA_VERSION = "1.0"

app = FastAPI(title="Guru API", version="0.4.0")
api_router = APIRouter(prefix="/api")


# ---------- Models (Schema v1.0 — FROZEN) ----------

SourceType = Literal["rozetka", "epicentr"]


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _short_hash(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8")).hexdigest()[:10]


def _build_product_key(source: str, source_id: Optional[str], url: Optional[str], title: Optional[str]) -> str:
    if source_id:
        return f"{source}:{source_id}"
    if url:
        try:
            return f"{source}:{_short_hash(urlparse(url).path)}"
        except Exception:
            pass
    return f"{source}:{_short_hash(title or '')}"


class ProductIn(BaseModel):
    model_config = ConfigDict(extra="ignore")

    schema_version: str = SCHEMA_VERSION
    source: SourceType
    source_id: Optional[str] = None
    external_id: Optional[str] = None  # alias for backward-compat with extension v0.3
    title: str
    brand: Optional[str] = None
    price: Optional[float] = None
    old_price: Optional[float] = None
    currency: str = "UAH"
    availability: Optional[bool] = None
    seller: Optional[str] = None
    seller_id: Optional[str] = None
    rating: Optional[float] = None
    reviews_count: Optional[int] = None
    saves_count: Optional[int] = None
    category: Optional[str] = None
    category_path: List[str] = Field(default_factory=list)
    badges: List[str] = Field(default_factory=list)
    image: Optional[str] = None
    images: List[str] = Field(default_factory=list)
    product_url: Optional[str] = None
    url: Optional[str] = None  # canonical alias for product_url
    position: Optional[int] = None
    listing_url: Optional[str] = None
    raw_data: Optional[Dict[str, Any]] = None
    product_key: Optional[str] = None  # accepted from client; recomputed server-side


class IngestProductsRequest(BaseModel):
    items: List[ProductIn]
    listing_url: Optional[str] = None
    listing_title: Optional[str] = None


class PageSnapshotIn(BaseModel):
    model_config = ConfigDict(extra="ignore")

    source: SourceType
    url: str
    title: Optional[str] = None
    category: Optional[str] = None
    category_path: List[str] = Field(default_factory=list)
    products_count: int = 0
    avg_price: Optional[float] = None
    min_price: Optional[float] = None
    max_price: Optional[float] = None
    badge_distribution: Dict[str, int] = Field(default_factory=dict)
    seller_distribution: Dict[str, int] = Field(default_factory=dict)


class CfSessionIn(BaseModel):
    user_agent: str
    cf_clearance: Optional[str] = None
    cf_bm: Optional[str] = None
    extra_cookies: Optional[Dict[str, str]] = None


# ---------- Auth ----------

async def require_extension_key(x_extension_api_key: Optional[str] = Header(default=None)):
    if x_extension_api_key != EXTENSION_API_KEY:
        raise HTTPException(status_code=401, detail="Invalid extension API key")
    return True


# ---------- Helpers ----------

def _calc_discount(price: Optional[float], old_price: Optional[float]) -> Optional[float]:
    if price is None or old_price is None or old_price <= 0 or price >= old_price:
        return None
    return round((1 - price / old_price) * 100, 1)


def _normalize_product(p: ProductIn) -> Dict[str, Any]:
    doc = p.model_dump()
    # Canonicalize fields
    sid = doc.get("source_id") or doc.get("external_id")
    doc["source_id"] = sid
    doc["external_id"] = sid  # keep both for backward-compat readers
    url = doc.get("url") or doc.get("product_url")
    doc["url"] = url
    doc["product_url"] = url
    if not doc.get("images") and doc.get("image"):
        doc["images"] = [doc["image"]]
    if not doc.get("image") and doc.get("images"):
        doc["image"] = doc["images"][0]
    doc["schema_version"] = SCHEMA_VERSION
    doc["product_key"] = _build_product_key(doc["source"], sid, url, doc.get("title"))
    doc["id"] = str(uuid.uuid4())
    doc["captured_at"] = now_iso()
    doc["discount_percent"] = _calc_discount(doc.get("price"), doc.get("old_price"))
    return doc


def _serialize_doc(doc: Dict[str, Any]) -> Dict[str, Any]:
    doc.pop("_id", None)
    return doc


async def _record_price_snapshots(docs: List[Dict[str, Any]]) -> None:
    """Always-on price history: append a thin snapshot doc per ingest with price > 0.

    `price_snapshots` is append-only; it powers `/api/products/by-key/{key}/history`
    and `/api/analytics/price-changes`.
    """
    snaps = []
    for d in docs:
        if d.get("price") is None:
            continue
        snaps.append(
            {
                "id": str(uuid.uuid4()),
                "product_key": d["product_key"],
                "source": d["source"],
                "source_id": d.get("source_id"),
                "url": d.get("url"),
                "title": d.get("title"),
                "price": d["price"],
                "old_price": d.get("old_price"),
                "discount_percent": d.get("discount_percent"),
                "availability": d.get("availability"),
                "seller": d.get("seller"),
                "image": d.get("image"),
                "captured_at": d["captured_at"],
            }
        )
    if snaps:
        try:
            await db.price_snapshots.insert_many(snaps, ordered=False)
        except Exception as e:
            logging.getLogger("guru").exception("price_snapshots insert failed: %s", e)


async def _stamp_lifecycle_dates(docs: List[Dict[str, Any]]) -> None:
    """Maintain first_seen_at / last_seen_at on `products` docs (Task 1 lifecycle layer).

    For each ingested doc:
      - first_seen_at = MIN(existing first_seen, current captured_at)
      - last_seen_at  = current captured_at (always overwritten)
    Done via a tiny update_many keyed by product_key (cheap with our index).
    """
    now_ts = now_iso()
    for d in docs:
        pk = d.get("product_key")
        if not pk:
            continue
        # Find any earlier first_seen_at for this product_key
        prior = await db.products.find_one(
            {"product_key": pk, "first_seen_at": {"$ne": None}},
            {"_id": 0, "first_seen_at": 1},
            sort=[("first_seen_at", 1)],
        )
        first_seen = (prior or {}).get("first_seen_at") or d["captured_at"]
        await db.products.update_many(
            {"product_key": pk},
            {"$set": {"first_seen_at": first_seen, "last_seen_at": now_ts}},
        )
        d["first_seen_at"] = first_seen
        d["last_seen_at"] = now_ts


async def _emit_market_signals(docs: List[Dict[str, Any]]) -> None:
    """For each ingested doc, compute signals against its full history and persist them.

    Idempotent per (kind, product_key, day) via `dedup_key` unique index.
    """
    if not docs:
        return
    all_sigs: List[Dict[str, Any]] = []
    for d in docs:
        pk = d.get("product_key")
        if not pk:
            continue
        # Full history (ASC) — includes the row we just inserted in price_snapshots.
        cur = (
            db.price_snapshots.find({"product_key": pk}, {"_id": 0})
            .sort("captured_at", 1)
        )
        history = await cur.to_list(2000)
        is_first_seen = len(history) <= 1
        last_before = None
        if len(history) >= 2:
            try:
                last_before = datetime.fromisoformat(
                    history[-2].get("captured_at").replace("Z", "+00:00")
                )
            except Exception:
                last_before = None
        sigs = build_signals_for_ingest(
            d, history, is_first_seen=is_first_seen, last_capture_before_now=last_before
        )
        all_sigs.extend(sigs)

    if not all_sigs:
        return
    # Upsert by dedup_key for idempotency.
    for s in all_sigs:
        try:
            await db.market_signals.update_one(
                {"dedup_key": s["dedup_key"]},
                {"$setOnInsert": s},
                upsert=True,
            )
        except Exception as e:
            logging.getLogger("guru.signals").warning("signal upsert failed: %s", e)


# ---------- Routes ----------

@api_router.get("/")
async def root():
    return {"app": "Guru", "status": "ok", "schema": SCHEMA_VERSION, "ts": now_iso()}


@api_router.get("/extension/health")
async def extension_health(_: bool = Depends(require_extension_key)):
    return {"ok": True, "server_time": now_iso(), "schema": SCHEMA_VERSION}


@api_router.post("/ingest/product")
async def ingest_product(payload: ProductIn, _: bool = Depends(require_extension_key)):
    doc = _normalize_product(payload)
    await db.products.insert_one(dict(doc))
    await _record_price_snapshots([doc])
    await _stamp_lifecycle_dates([doc])
    await _emit_market_signals([doc])
    return {"ok": True, "id": doc["id"], "product_key": doc["product_key"]}


@api_router.post("/ingest/products")
async def ingest_products(payload: IngestProductsRequest, _: bool = Depends(require_extension_key)):
    if not payload.items:
        return {"ok": True, "inserted": 0}
    docs = [_normalize_product(p) for p in payload.items]
    for d in docs:
        if payload.listing_url and not d.get("listing_url"):
            d["listing_url"] = payload.listing_url
    try:
        await db.products.insert_many([dict(d) for d in docs], ordered=False)
    except Exception as e:
        logging.getLogger("guru").warning("products insert_many partial: %s", e)
    await _record_price_snapshots(docs)
    await _stamp_lifecycle_dates(docs)
    await _emit_market_signals(docs)
    return {
        "ok": True,
        "inserted": len(docs),
        "ids": [d["id"] for d in docs],
        "product_keys": [d["product_key"] for d in docs],
    }


@api_router.post("/ingest/page")
async def ingest_page(payload: PageSnapshotIn, _: bool = Depends(require_extension_key)):
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["captured_at"] = now_iso()
    await db.pages.insert_one(dict(doc))
    return {"ok": True, "id": doc["id"]}


@api_router.post("/ingest/cf-session")
async def ingest_cf_session(payload: CfSessionIn, _: bool = Depends(require_extension_key)):
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["captured_at"] = now_iso()
    await db.cf_sessions.insert_one(dict(doc))
    return {"ok": True, "id": doc["id"]}


# ---------- Public read endpoints ----------

@api_router.get("/products")
async def list_products(
    source: Optional[SourceType] = None,
    search: Optional[str] = None,
    seller: Optional[str] = None,
    limit: int = Query(default=50, ge=1, le=500),
    skip: int = Query(default=0, ge=0),
):
    q: Dict[str, Any] = {}
    if source:
        q["source"] = source
    if seller:
        q["seller"] = seller
    if search:
        q["title"] = {"$regex": search, "$options": "i"}

    cursor = (
        db.products.find(q, {"_id": 0})
        .sort("captured_at", -1)
        .skip(skip)
        .limit(limit)
    )
    items = [_serialize_doc(d) for d in await cursor.to_list(limit)]
    total = await db.products.count_documents(q)
    return {"items": items, "total": total, "limit": limit, "skip": skip}


@api_router.get("/products/by-key/{product_key}")
async def product_by_key(product_key: str):
    doc = await db.products.find_one({"product_key": product_key}, {"_id": 0}, sort=[("captured_at", -1)])
    if not doc:
        raise HTTPException(status_code=404, detail="Product not found")
    return doc


@api_router.get("/products/by-key/{product_key}/history")
async def product_history(product_key: str, limit: int = Query(default=200, ge=1, le=2000)):
    cursor = (
        db.price_snapshots.find({"product_key": product_key}, {"_id": 0})
        .sort("captured_at", 1)
        .limit(limit)
    )
    items = await cursor.to_list(limit)
    return {"product_key": product_key, "count": len(items), "items": items}


@api_router.get("/products/by-key/{product_key}/insights")
async def product_insights(product_key: str):
    """Rule-based market insights for a single product (no AI).

    Insights:
      - new_minimum: current price equals historical minimum
      - consecutive_drops: price dropped N captures in a row
      - above_avg_discount: discount_percent above category average
      - price_stable: price unchanged for N captures
      - high_engagement: many reviews
      - long_tracked: product has > N snapshots (popular to collect)
    """
    cursor = (
        db.price_snapshots.find({"product_key": product_key}, {"_id": 0})
        .sort("captured_at", 1)
    )
    history = await cursor.to_list(2000)
    if not history:
        raise HTTPException(status_code=404, detail="No history for product")
    latest_doc = await db.products.find_one(
        {"product_key": product_key}, {"_id": 0}, sort=[("captured_at", -1)]
    )

    insights: List[Dict[str, Any]] = []

    prices = [h["price"] for h in history if h.get("price") is not None]
    if prices:
        last = prices[-1]
        min_p = min(prices)
        max_p = max(prices)

        if last == min_p and len(prices) >= 2:
            insights.append(
                {
                    "kind": "new_minimum",
                    "level": "info",
                    "icon": "trending-down",
                    "title": "Нова мінімальна ціна",
                    "detail": f"Поточна {round(last)} ₴ — найнижча за {len(prices)} спостережень",
                }
            )
        elif last == max_p and len(prices) >= 3:
            insights.append(
                {
                    "kind": "new_maximum",
                    "level": "warn",
                    "icon": "trending-up",
                    "title": "Нова максимальна ціна",
                    "detail": f"Поточна {round(last)} ₴ — найвища за {len(prices)} спостережень",
                }
            )

        # consecutive drops/rises
        cons_drops = 0
        cons_rises = 0
        for i in range(len(prices) - 1, 0, -1):
            if prices[i] < prices[i - 1]:
                cons_drops += 1
            else:
                break
        for i in range(len(prices) - 1, 0, -1):
            if prices[i] > prices[i - 1]:
                cons_rises += 1
            else:
                break
        if cons_drops >= 2:
            insights.append(
                {
                    "kind": "consecutive_drops",
                    "level": "good",
                    "icon": "arrow-down-right",
                    "title": f"Ціна падає {cons_drops} оновлень поспіль",
                    "detail": "Тренд на зниження",
                }
            )
        elif cons_rises >= 2:
            insights.append(
                {
                    "kind": "consecutive_rises",
                    "level": "warn",
                    "icon": "arrow-up-right",
                    "title": f"Ціна росте {cons_rises} оновлень поспіль",
                    "detail": "Тренд на підвищення",
                }
            )

        # stability
        if len(prices) >= 4 and max_p - min_p < max_p * 0.01:
            insights.append(
                {
                    "kind": "price_stable",
                    "level": "neutral",
                    "icon": "minus",
                    "title": "Ціна стабільна",
                    "detail": f"Не змінювалась за {len(prices)} спостережень",
                }
            )

    # discount above category average
    if latest_doc:
        dp = latest_doc.get("discount_percent")
        cat = latest_doc.get("category_path") or []
        cat_root = cat[0] if cat else latest_doc.get("category")
        if dp and cat_root:
            cursor = db.products.aggregate(
                [
                    {
                        "$match": {
                            "discount_percent": {"$gt": 0},
                            "$or": [
                                {"category": cat_root},
                                {"category_path": cat_root},
                            ],
                        }
                    },
                    {"$group": {"_id": None, "avg": {"$avg": "$discount_percent"}, "n": {"$sum": 1}}},
                ]
            )
            async for row in cursor:
                if row.get("n", 0) >= 2 and row.get("avg") and dp > row["avg"] * 1.2:
                    insights.append(
                        {
                            "kind": "above_avg_discount",
                            "level": "good",
                            "icon": "tag",
                            "title": f"Знижка вища за середню по «{cat_root}»",
                            "detail": f"−{round(dp)}% vs середня −{round(row['avg'])}%",
                        }
                    )

        # engagement
        rc = latest_doc.get("reviews_count") or 0
        if rc >= 100:
            insights.append(
                {
                    "kind": "high_engagement",
                    "level": "info",
                    "icon": "flame",
                    "title": "Високий інтерес покупців",
                    "detail": f"{rc} відгуків",
                }
            )

        # availability flip
        avail_seq = [h.get("availability") for h in history if h.get("availability") is not None]
        if len(avail_seq) >= 2 and avail_seq[-1] is False and avail_seq[-2] is True:
            insights.append(
                {
                    "kind": "out_of_stock",
                    "level": "warn",
                    "icon": "shield-alert",
                    "title": "Товар закінчився",
                    "detail": "Перейшов у статус «немає в наявності»",
                }
            )

    if len(history) >= 10:
        insights.append(
            {
                "kind": "long_tracked",
                "level": "neutral",
                "icon": "clock",
                "title": "Активно відстежується",
                "detail": f"{len(history)} захоплень в історії",
            }
        )

    return {"product_key": product_key, "count": len(insights), "items": insights}


@api_router.get("/analytics/summary")
async def analytics_summary():
    total = await db.products.count_documents({})
    rozetka_count = await db.products.count_documents({"source": "rozetka"})
    epicentr_count = await db.products.count_documents({"source": "epicentr"})

    pipeline_avg = [
        {"$match": {"price": {"$ne": None}}},
        {
            "$group": {
                "_id": "$source",
                "avg_price": {"$avg": "$price"},
                "min_price": {"$min": "$price"},
                "max_price": {"$max": "$price"},
                "count": {"$sum": 1},
            }
        },
    ]
    by_source = {}
    async for row in db.products.aggregate(pipeline_avg):
        by_source[row["_id"]] = {
            "avg_price": round(row["avg_price"], 2) if row.get("avg_price") else None,
            "min_price": row.get("min_price"),
            "max_price": row.get("max_price"),
            "count": row["count"],
        }

    pipeline_badges = [
        {"$unwind": "$badges"},
        {"$group": {"_id": "$badges", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    badge_distribution = {}
    async for row in db.products.aggregate(pipeline_badges):
        badge_distribution[row["_id"] or "none"] = row["count"]

    yesterday = (datetime.now(timezone.utc) - timedelta(days=1)).isoformat()
    last_24h = await db.products.count_documents({"captured_at": {"$gte": yesterday}})
    pages_count = await db.pages.count_documents({})
    snapshots_count = await db.price_snapshots.count_documents({})
    unique_products = len(
        await db.price_snapshots.distinct("product_key")
    ) if snapshots_count else 0

    # total unique sellers + avg discount across products with discount
    total_sellers = len(await db.products.distinct("seller", {"seller": {"$ne": None}}))
    avg_discount_overall = None
    cursor = db.products.aggregate(
        [
            {"$match": {"discount_percent": {"$gt": 0}}},
            {"$group": {"_id": None, "avg": {"$avg": "$discount_percent"}}},
        ]
    )
    async for row in cursor:
        avg_discount_overall = round(row["avg"], 1) if row.get("avg") else None

    return {
        "total_products": total,
        "rozetka_count": rozetka_count,
        "epicentr_count": epicentr_count,
        "captured_last_24h": last_24h,
        "pages_collected": pages_count,
        "price_snapshots": snapshots_count,
        "tracked_products": unique_products,
        "total_sellers": total_sellers,
        "avg_discount_overall": avg_discount_overall,
        "by_source": by_source,
        "badge_distribution": badge_distribution,
        "schema_version": SCHEMA_VERSION,
        "ts": now_iso(),
    }


@api_router.get("/analytics/today")
async def analytics_today():
    """Live data-loop counters: what came through the system in the last 24h."""
    yesterday = (datetime.now(timezone.utc) - timedelta(days=1)).isoformat()

    products_24h = await db.products.count_documents({"captured_at": {"$gte": yesterday}})
    snapshots_24h = await db.price_snapshots.count_documents({"captured_at": {"$gte": yesterday}})
    pages_24h = await db.pages.count_documents({"captured_at": {"$gte": yesterday}})

    # price-changes 24h (products with >=2 snapshots in window)
    pipeline_changes = [
        {"$match": {"captured_at": {"$gte": yesterday}, "price": {"$ne": None}}},
        {"$group": {"_id": "$product_key", "n": {"$sum": 1}, "min": {"$min": "$price"}, "max": {"$max": "$price"}}},
        {"$match": {"n": {"$gte": 2}, "$expr": {"$ne": ["$min", "$max"]}}},
        {"$count": "changed"},
    ]
    price_changes_24h = 0
    async for r in db.price_snapshots.aggregate(pipeline_changes):
        price_changes_24h = r.get("changed", 0)

    # signals 24h (drops >=5%)
    pipeline_signals = [
        {"$match": {"captured_at": {"$gte": yesterday}, "price": {"$ne": None}}},
        {"$sort": {"captured_at": 1}},
        {
            "$group": {
                "_id": "$product_key",
                "first": {"$first": "$price"},
                "last": {"$last": "$price"},
                "n": {"$sum": 1},
            }
        },
        {"$match": {"n": {"$gte": 2}}},
        {
            "$addFields": {
                "pct": {
                    "$cond": [
                        {"$gt": ["$first", 0]},
                        {"$multiply": [{"$divide": [{"$subtract": ["$last", "$first"]}, "$first"]}, 100]},
                        None,
                    ]
                }
            }
        },
        {"$match": {"pct": {"$lte": -5}}},
        {"$count": "signals"},
    ]
    signals_24h = 0
    async for r in db.price_snapshots.aggregate(pipeline_signals):
        signals_24h = r.get("signals", 0)

    sellers_24h = len(
        await db.products.distinct(
            "seller", {"captured_at": {"$gte": yesterday}, "seller": {"$ne": None}}
        )
    )
    sources_active = await db.products.distinct(
        "source", {"captured_at": {"$gte": yesterday}}
    )

    # Last activity
    last_capture = await db.products.find_one(
        {}, {"_id": 0, "captured_at": 1}, sort=[("captured_at", -1)]
    )
    last_capture_at = last_capture.get("captured_at") if last_capture else None

    return {
        "products_24h": products_24h,
        "snapshots_24h": snapshots_24h,
        "pages_24h": pages_24h,
        "price_changes_24h": price_changes_24h,
        "signals_24h": signals_24h,
        "sellers_24h": sellers_24h,
        "sources_active": sources_active,
        "last_capture_at": last_capture_at,
        "ts": now_iso(),
    }


@api_router.get("/analytics/top-sellers")
async def analytics_top_sellers(limit: int = 10, source: Optional[SourceType] = None):
    match: Dict[str, Any] = {"seller": {"$ne": None}}
    if source:
        match["source"] = source
    pipeline = [
        {"$match": match},
        {
            "$group": {
                "_id": "$seller",
                "products": {"$sum": 1},
                "avg_price": {"$avg": "$price"},
                "sources": {"$addToSet": "$source"},
            }
        },
        {"$sort": {"products": -1}},
        {"$limit": limit},
    ]
    out = []
    async for row in db.products.aggregate(pipeline):
        out.append(
            {
                "seller": row["_id"],
                "products": row["products"],
                "avg_price": round(row["avg_price"], 2) if row.get("avg_price") else None,
                "sources": row.get("sources", []),
            }
        )
    return {"items": out}


@api_router.get("/analytics/source-distribution")
async def analytics_source_distribution():
    pipeline = [
        {"$unwind": {"path": "$badges", "preserveNullAndEmptyArrays": True}},
        {"$group": {"_id": "$badges", "count": {"$sum": 1}}},
    ]
    raw = {}
    async for row in db.products.aggregate(pipeline):
        key = row["_id"] or "other"
        raw[key] = row["count"]

    canonical = {
        "Топ продажів": raw.get("top_sales", 0),
        "Реклама": raw.get("ad", 0) + raw.get("promo", 0),
        "Rozetka власні": await db.products.count_documents(
            {"source": "rozetka", "seller": {"$in": [None, "Rozetka", "ROZETKA"]}}
        ),
        "Фулфілмент": raw.get("fulfilled", 0),
        "Інші": 0,
    }
    total = await db.products.count_documents({})
    others = total - sum(v for k, v in canonical.items() if k != "Інші")
    canonical["Інші"] = max(others, 0)
    return {"items": [{"label": k, "value": v} for k, v in canonical.items()], "total": total}


# ---------- Snapshot history analytics ----------

def _parse_window(window: str) -> timedelta:
    window = (window or "7d").strip().lower()
    if window.endswith("h"):
        return timedelta(hours=int(window[:-1] or 24))
    if window.endswith("d"):
        return timedelta(days=int(window[:-1] or 7))
    return timedelta(days=7)


@api_router.get("/analytics/price-changes")
async def analytics_price_changes(
    window: str = "7d",
    direction: Literal["down", "up", "any"] = "down",
    source: Optional[SourceType] = None,
    limit: int = Query(default=20, ge=1, le=100),
):
    """Top products by price change within the time window.

    Compares the earliest snapshot in the window vs the latest snapshot.
    Returns absolute and percentage delta.
    """
    delta = _parse_window(window)
    since = (datetime.now(timezone.utc) - delta).isoformat()

    match: Dict[str, Any] = {"captured_at": {"$gte": since}, "price": {"$ne": None}}
    if source:
        match["source"] = source

    pipeline = [
        {"$match": match},
        {"$sort": {"captured_at": 1}},
        {
            "$group": {
                "_id": "$product_key",
                "first": {"$first": "$$ROOT"},
                "last": {"$last": "$$ROOT"},
                "snapshots": {"$sum": 1},
                "min_price": {"$min": "$price"},
                "max_price": {"$max": "$price"},
            }
        },
        {"$match": {"snapshots": {"$gte": 2}}},
        {
            "$addFields": {
                "delta_abs": {"$subtract": ["$last.price", "$first.price"]},
            }
        },
        {
            "$addFields": {
                "delta_pct": {
                    "$cond": [
                        {"$gt": ["$first.price", 0]},
                        {"$multiply": [{"$divide": ["$delta_abs", "$first.price"]}, 100]},
                        None,
                    ]
                }
            }
        },
    ]
    if direction == "down":
        pipeline.append({"$match": {"delta_abs": {"$lt": 0}}})
        pipeline.append({"$sort": {"delta_pct": 1}})
    elif direction == "up":
        pipeline.append({"$match": {"delta_abs": {"$gt": 0}}})
        pipeline.append({"$sort": {"delta_pct": -1}})
    else:
        pipeline.append({"$match": {"delta_abs": {"$ne": 0}}})
        pipeline.append({"$sort": {"delta_pct": -1}})
    pipeline.append({"$limit": limit})

    items = []
    async for row in db.price_snapshots.aggregate(pipeline):
        first = row.get("first", {})
        last = row.get("last", {})
        items.append(
            {
                "product_key": row["_id"],
                "source": last.get("source") or first.get("source"),
                "title": last.get("title") or first.get("title"),
                "url": last.get("url") or first.get("url"),
                "image": last.get("image") or first.get("image"),
                "seller": last.get("seller"),
                "first_price": first.get("price"),
                "last_price": last.get("price"),
                "min_price": row.get("min_price"),
                "max_price": row.get("max_price"),
                "delta_abs": round(row.get("delta_abs") or 0, 2),
                "delta_pct": round(row.get("delta_pct") or 0, 2),
                "snapshots": row.get("snapshots"),
                "first_captured": first.get("captured_at"),
                "last_captured": last.get("captured_at"),
            }
        )
    return {"window": window, "direction": direction, "items": items, "count": len(items)}


@api_router.get("/analytics/top-discounts")
async def analytics_top_discounts(
    source: Optional[SourceType] = None,
    limit: int = Query(default=20, ge=1, le=100),
):
    """Latest snapshot of each product with the largest discount_percent."""
    match: Dict[str, Any] = {"discount_percent": {"$ne": None, "$gt": 0}}
    if source:
        match["source"] = source
    pipeline = [
        {"$match": match},
        {"$sort": {"captured_at": -1}},
        {
            "$group": {
                "_id": "$product_key",
                "latest": {"$first": "$$ROOT"},
            }
        },
        {"$replaceRoot": {"newRoot": "$latest"}},
        {"$sort": {"discount_percent": -1}},
        {"$limit": limit},
        {
            "$project": {
                "_id": 0,
                "product_key": 1,
                "source": 1,
                "title": 1,
                "url": 1,
                "image": 1,
                "seller": 1,
                "price": 1,
                "old_price": 1,
                "discount_percent": 1,
                "captured_at": 1,
            }
        },
    ]
    items = []
    async for row in db.products.aggregate(pipeline):
        items.append(row)
    return {"items": items, "count": len(items)}


@api_router.get("/analytics/seller-dominance")
async def analytics_seller_dominance(
    source: Optional[SourceType] = None,
    limit: int = Query(default=10, ge=1, le=50),
):
    """Seller share of unique products tracked."""
    match: Dict[str, Any] = {"seller": {"$ne": None}}
    if source:
        match["source"] = source
    pipeline = [
        {"$match": match},
        {
            "$group": {
                "_id": {"product_key": "$product_key", "seller": "$seller"},
                "snapshots": {"$sum": 1},
            }
        },
        {
            "$group": {
                "_id": "$_id.seller",
                "unique_products": {"$sum": 1},
                "total_snapshots": {"$sum": "$snapshots"},
            }
        },
        {"$sort": {"unique_products": -1}},
        {"$limit": limit},
    ]
    rows = []
    total = 0
    async for row in db.products.aggregate(pipeline):
        total += row["unique_products"]
        rows.append(
            {
                "seller": row["_id"],
                "unique_products": row["unique_products"],
                "total_snapshots": row["total_snapshots"],
            }
        )
    for r in rows:
        r["share_pct"] = round((r["unique_products"] / total) * 100, 1) if total else 0
    return {"items": rows, "total_unique_products": total}


@api_router.get("/pages")
async def list_pages(limit: int = 30):
    cursor = db.pages.find({}, {"_id": 0}).sort("captured_at", -1).limit(limit)
    items = await cursor.to_list(limit)
    return {"items": items, "count": len(items)}


# ---------- Seller analytics (full) ----------

@api_router.get("/analytics/sellers")
async def analytics_sellers(
    source: Optional[SourceType] = None,
    limit: int = Query(default=50, ge=1, le=500),
    sort: Literal["products", "avg_price", "avg_discount", "share"] = "products",
):
    """Full seller analytics: unique products, avg price, avg discount, market share, sources, last seen."""
    match: Dict[str, Any] = {"seller": {"$ne": None}}
    if source:
        match["source"] = source

    # Step 1: per-seller per-product latest snapshot (avoid double-counting)
    pipeline = [
        {"$match": match},
        {"$sort": {"captured_at": -1}},
        {
            "$group": {
                "_id": {"product_key": "$product_key", "seller": "$seller"},
                "latest": {"$first": "$$ROOT"},
            }
        },
        {
            "$group": {
                "_id": "$_id.seller",
                "unique_products": {"$sum": 1},
                "avg_price": {"$avg": "$latest.price"},
                "avg_discount": {"$avg": "$latest.discount_percent"},
                "with_discount": {
                    "$sum": {
                        "$cond": [
                            {"$gt": ["$latest.discount_percent", 0]},
                            1,
                            0,
                        ]
                    }
                },
                "available": {
                    "$sum": {
                        "$cond": [{"$eq": ["$latest.availability", True]}, 1, 0]
                    }
                },
                "sources": {"$addToSet": "$latest.source"},
                "last_seen": {"$max": "$latest.captured_at"},
                "min_price": {"$min": "$latest.price"},
                "max_price": {"$max": "$latest.price"},
            }
        },
    ]
    rows: List[Dict[str, Any]] = []
    total_unique = 0
    async for row in db.products.aggregate(pipeline):
        total_unique += row.get("unique_products", 0) or 0
        rows.append(
            {
                "seller": row["_id"],
                "unique_products": row.get("unique_products", 0),
                "avg_price": round(row["avg_price"], 2) if row.get("avg_price") else None,
                "avg_discount": round(row["avg_discount"], 1) if row.get("avg_discount") else None,
                "with_discount": row.get("with_discount", 0),
                "available": row.get("available", 0),
                "sources": row.get("sources", []),
                "last_seen": row.get("last_seen"),
                "min_price": row.get("min_price"),
                "max_price": row.get("max_price"),
            }
        )
    for r in rows:
        r["share_pct"] = (
            round((r["unique_products"] / total_unique) * 100, 1) if total_unique else 0
        )

    sort_key_map = {
        "products": ("unique_products", -1),
        "avg_price": ("avg_price", -1),
        "avg_discount": ("avg_discount", -1),
        "share": ("share_pct", -1),
    }
    key, direction = sort_key_map[sort]
    rows.sort(key=lambda r: (r.get(key) or 0), reverse=(direction == -1))
    return {
        "items": rows[:limit],
        "total_sellers": len(rows),
        "total_unique_products": total_unique,
    }


@api_router.get("/analytics/price-signals")
async def analytics_price_signals(
    window: str = "24h",
    min_drop_pct: float = Query(default=5.0, ge=0.0, le=100.0),
    source: Optional[SourceType] = None,
    limit: int = Query(default=20, ge=1, le=200),
):
    """Recent price-drop signals: products that dropped >= min_drop_pct in the window."""
    delta = _parse_window(window)
    since = (datetime.now(timezone.utc) - delta).isoformat()
    match: Dict[str, Any] = {"captured_at": {"$gte": since}, "price": {"$ne": None}}
    if source:
        match["source"] = source

    pipeline = [
        {"$match": match},
        {"$sort": {"captured_at": 1}},
        {
            "$group": {
                "_id": "$product_key",
                "first": {"$first": "$$ROOT"},
                "last": {"$last": "$$ROOT"},
                "snapshots": {"$sum": 1},
            }
        },
        {"$match": {"snapshots": {"$gte": 2}}},
        {
            "$addFields": {
                "delta_abs": {"$subtract": ["$last.price", "$first.price"]},
            }
        },
        {
            "$addFields": {
                "delta_pct": {
                    "$cond": [
                        {"$gt": ["$first.price", 0]},
                        {"$multiply": [{"$divide": ["$delta_abs", "$first.price"]}, 100]},
                        None,
                    ]
                }
            }
        },
        {"$match": {"delta_pct": {"$lte": -min_drop_pct}}},
        {"$sort": {"delta_pct": 1, "last.captured_at": -1}},
        {"$limit": limit},
    ]
    items = []
    async for row in db.price_snapshots.aggregate(pipeline):
        first = row.get("first", {})
        last = row.get("last", {})
        items.append(
            {
                "product_key": row["_id"],
                "source": last.get("source"),
                "title": last.get("title"),
                "url": last.get("url"),
                "image": last.get("image"),
                "seller": last.get("seller"),
                "first_price": first.get("price"),
                "last_price": last.get("price"),
                "delta_abs": round(row.get("delta_abs") or 0, 2),
                "delta_pct": round(row.get("delta_pct") or 0, 2),
                "first_captured": first.get("captured_at"),
                "last_captured": last.get("captured_at"),
            }
        )
    return {
        "window": window,
        "min_drop_pct": min_drop_pct,
        "items": items,
        "count": len(items),
    }


# ---------- Export ----------

EXPORT_FIELDS = [
    "captured_at",
    "source",
    "product_key",
    "source_id",
    "title",
    "brand",
    "category",
    "price",
    "old_price",
    "discount_percent",
    "currency",
    "availability",
    "seller",
    "rating",
    "reviews_count",
    "badges",
    "url",
    "image",
    "position",
    "listing_url",
]


async def _query_for_export(
    source: Optional[str],
    seller: Optional[str],
    category: Optional[str],
    search: Optional[str],
    min_price: Optional[float],
    max_price: Optional[float],
    min_discount: Optional[float],
    date_from: Optional[str],
    date_to: Optional[str],
    limit: int,
) -> List[Dict[str, Any]]:
    q: Dict[str, Any] = {}
    if source:
        q["source"] = source
    if seller:
        q["seller"] = seller
    if category:
        q["category_path"] = {"$regex": category, "$options": "i"}
    if search:
        q["title"] = {"$regex": search, "$options": "i"}
    price_q: Dict[str, Any] = {}
    if min_price is not None:
        price_q["$gte"] = min_price
    if max_price is not None:
        price_q["$lte"] = max_price
    if price_q:
        q["price"] = price_q
    if min_discount is not None:
        q["discount_percent"] = {"$gte": min_discount}
    captured_q: Dict[str, Any] = {}
    if date_from:
        captured_q["$gte"] = date_from
    if date_to:
        captured_q["$lte"] = date_to
    if captured_q:
        q["captured_at"] = captured_q

    cursor = db.products.find(q, {"_id": 0}).sort("captured_at", -1).limit(limit)
    return await cursor.to_list(limit)


def _row_for_export(d: Dict[str, Any]) -> List[Any]:
    out = []
    for f in EXPORT_FIELDS:
        v = d.get(f)
        if isinstance(v, list):
            v = ", ".join(str(x) for x in v)
        out.append(v if v is not None else "")
    return out


@api_router.get("/export/products.csv")
async def export_products_csv(
    source: Optional[SourceType] = None,
    seller: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    min_price: Optional[float] = None,
    max_price: Optional[float] = None,
    min_discount: Optional[float] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = Query(default=10000, ge=1, le=100000),
):
    rows = await _query_for_export(
        source, seller, category, search,
        min_price, max_price, min_discount, date_from, date_to, limit,
    )
    buf = io.StringIO()
    buf.write("\ufeff")  # BOM for Excel UTF-8 compatibility
    writer = csv.writer(buf, delimiter=",", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(EXPORT_FIELDS)
    for d in rows:
        writer.writerow(_row_for_export(d))
    data = buf.getvalue().encode("utf-8")
    fname = f"guru-products-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.csv"
    return StreamingResponse(
        iter([data]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api_router.get("/export/products.xlsx")
async def export_products_xlsx(
    source: Optional[SourceType] = None,
    seller: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    min_price: Optional[float] = None,
    max_price: Optional[float] = None,
    min_discount: Optional[float] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = Query(default=10000, ge=1, le=100000),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    rows = await _query_for_export(
        source, seller, category, search,
        min_price, max_price, min_discount, date_from, date_to, limit,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Guru Products"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0A0A0A")
    ws.append(EXPORT_FIELDS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for d in rows:
        ws.append(_row_for_export(d))
    # Column widths heuristic
    widths = {
        "captured_at": 22, "source": 10, "product_key": 22, "source_id": 14,
        "title": 60, "brand": 16, "category": 28, "price": 10, "old_price": 10,
        "discount_percent": 10, "currency": 8, "availability": 12, "seller": 24,
        "rating": 8, "reviews_count": 10, "badges": 24, "url": 60, "image": 60,
        "position": 8, "listing_url": 60,
    }
    for idx, f in enumerate(EXPORT_FIELDS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = widths.get(f, 16)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"guru-products-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------- Mount ----------

app.include_router(api_router)
# Task 1 — Foundation Analytics Layer (watchlist, lifecycle, signals, market overview v2, feed)
app.include_router(make_intelligence_router(db, require_extension_key))

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger("guru")


@app.on_event("startup")
async def ensure_indexes():
    """Idempotent index creation — critical for scaling."""
    try:
        await db.products.create_index("source")
        await db.products.create_index("captured_at")
        await db.products.create_index("seller")
        await db.products.create_index("product_key")
        await db.products.create_index("first_seen_at")
        await db.products.create_index("last_seen_at")
        await db.products.create_index([("title", "text")])
        await db.price_snapshots.create_index("product_key")
        await db.price_snapshots.create_index("captured_at")
        await db.price_snapshots.create_index([("product_key", 1), ("captured_at", 1)])
        await db.pages.create_index("captured_at")
        # Task 1 collections
        await db.market_signals.create_index("captured_at")
        await db.market_signals.create_index("kind")
        await db.market_signals.create_index("product_key")
        await db.market_signals.create_index("seller")
        await db.market_signals.create_index("category_path")
        await db.market_signals.create_index("dedup_key", unique=True)
        await db.watchlists.create_index("category_key")
        await db.watchlists.create_index("enabled")
        logger.info("Indexes ensured")
    except Exception as e:
        logger.warning("Index creation skipped: %s", e)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
